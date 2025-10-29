import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
from urllib.parse import quote
import time
import json
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
import logging
import subprocess
import os
import shutil
import tempfile
import uuid
import random
from typing import List, Dict, Optional, Tuple, Set
from selenium.common.exceptions import TimeoutException
import gc
from decimal import Decimal
from .autopiter_parser import get_next_proxy, make_request, get_brands_by_artikul, get_brands_by_artikul_emex, get_brands_by_artikul_armtek, _create_chrome_driver_robust

# Настройки для парсинга прайс-листа
TIMEOUT = 15  # Увеличиваем таймаут
SELENIUM_TIMEOUT = 20
PAGE_LOAD_TIMEOUT = 20
MAX_HTTP_RETRIES = 3  # Максимальное количество попыток HTTP-запросов

# Коды поставщиков для каждой площадки
SUPPLIER_CODES = {
    'autopiter': ['30399', '36364', '32994', '22771', '33305', '40251', '39479', '40112'],
    'emex': ['QFRD', 'QFRG'],
    'armtek': []  # Нет кода поставщика
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Cache-Control": "max-age=0",
    "Sec-Ch-Ua": '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
}

def log_debug(message):
    print(f"[DEBUG] {message}")

def _norm_brand(val: str) -> str:
    try:
        s = (val or '').strip().upper()
        return re.sub(r"[^0-9A-ZА-ЯЁ]+", "", s)
    except Exception:
        return ''

def parse_price_list_file(file_path: str) -> List[Dict]:
    """Парсит Excel файл прайс-листа и возвращает список позиций.
    Расширенный импорт: ищет строку заголовков в первых ~60 строках,
    поддерживает синонимы названий колонок и «шапки» с пустыми/мердж-ячейками.
    """
    try:
        # Читаем без заголовков с повышенной совместимостью
        df_raw = pd.read_excel(
            file_path,
            header=None,
            dtype=str,
            keep_default_na=False,
            engine='openpyxl'
        )

        def norm(val):
            try:
                return str(val).strip().lower()
            except Exception:
                return ''

        header_row = None
        man_idx = art_idx = None
        supp_idx = nom_idx = qty_idx = price_idx = None

        # Синонимы для поиска колонок
        manufacturer_keys = ['производител', 'бренд', 'марка', 'изготовител']
        article_keys = ['артикул', 'номер', 'код детали', 'код производителя']
        supplier_keys = ['код поставщика', 'код товара', 'код пост', 'поставщик', 'код склада']
        name_keys = ['номенклатура', 'наимен', 'описание', 'товар']
        qty_keys = ['колич', 'остаток', 'в наличии', 'наличие', 'остатки']
        price_keys = ['цена', 'стоим', 'оптов', 'рознич']

        def find_idx(row_vals, keys):
            try:
                return next((j for j, c in enumerate(row_vals) if any(k in c for k in keys)), None)
            except Exception:
                return None

        # Увеличиваем окно поиска шапки
        scan_rows = min(60, len(df_raw))
        for i in range(scan_rows):
            row_vals = [norm(v) for v in df_raw.iloc[i].tolist()]
            tmp_man = find_idx(row_vals, manufacturer_keys)
            tmp_art = find_idx(row_vals, article_keys)
            if tmp_man is not None and tmp_art is not None:
                header_row = i
                man_idx, art_idx = tmp_man, tmp_art
                supp_idx = find_idx(row_vals, supplier_keys)
                nom_idx = find_idx(row_vals, name_keys)
                qty_idx = find_idx(row_vals, qty_keys)
                price_idx = find_idx(row_vals, price_keys)
                break

        if header_row is None:
            raise ValueError('Не найдены обязательные колонки: Производитель и Артикул')

        headers = df_raw.iloc[header_row].tolist()
        for k in range(len(headers)):
            if pd.isna(headers[k]) or str(headers[k]).strip() == '':
                headers[k] = headers[k-1] if k > 0 else f'col_{k}'

        # Берём данные после шапки и чистим пустые строки
        df = df_raw.iloc[header_row + 1:].reset_index(drop=True)
        df.columns = headers
        df = df.dropna(how='all')

        supplier_col = df.columns[supp_idx] if supp_idx is not None and supp_idx < len(df.columns) else None
        manufacturer_col = df.columns[man_idx]
        article_col = df.columns[art_idx]
        nomenclature_col = df.columns[nom_idx] if nom_idx is not None and nom_idx < len(df.columns) else None
        quantity_col = df.columns[qty_idx] if qty_idx is not None and qty_idx < len(df.columns) else None
        price_col = df.columns[price_idx] if price_idx is not None and price_idx < len(df.columns) else None

        def to_int_safe(v):
            try:
                return int(float(str(v).replace(' ', '').replace('\xa0', '')))
            except Exception:
                return 0

        def to_float_safe(v):
            try:
                s = str(v).replace(' ', '').replace('\xa0', '').replace('₽', '').replace(',', '.')
                return float(s)
            except Exception:
                return None

        items: List[Dict] = []
        for index, row in df.iterrows():
            try:
                manufacturer = str(row[manufacturer_col]).strip() if pd.notna(row[manufacturer_col]) else ''
                article = str(row[article_col]).strip() if pd.notna(row[article_col]) else ''
                if not manufacturer or not article:
                    continue
                item = {
                    'supplier_code': (str(row[supplier_col]).strip() if supplier_col and pd.notna(row[supplier_col]) else ''),
                    'manufacturer': manufacturer,
                    'article': article,
                    'nomenclature': (str(row[nomenclature_col]).strip() if nomenclature_col and pd.notna(row[nomenclature_col]) else ''),
                    'quantity': (to_int_safe(row[quantity_col]) if quantity_col and pd.notna(row[quantity_col]) else 0),
                    'our_price': (to_float_safe(row[price_col]) if price_col and pd.notna(row[price_col]) else None),
                }
                items.append(item)
            except Exception as e:
                log_debug(f'Ошибка парсинга строки {index + 1}: {str(e)}')
                continue

        log_debug(f'Парсинг завершен. Найдено {len(items)} позиций')
        return items

    except Exception as e:
        log_debug(f'Ошибка парсинга файла: {str(e)}')
        return []

def check_autopiter_item(supplier_code: str, manufacturer: str, article: str, competitor_brand_filter: str = None) -> Dict:
    """Проверяет наличие позиции на АвтоПитере и анализирует цены"""
    print(f"[DEBUG] ===== НАЧАЛО check_autopiter_item =====")
    print(f"[DEBUG] Входные параметры: supplier_code={supplier_code}, manufacturer={manufacturer}, article={article}")
    
    result = {
        'is_found': False,
        'marketplace_price': None,
        'min_competitor_price': None,
        'competitor_brand': None,
        'quantity_in_stock': None,
        'competitor_quantity': None,  # Добавляем это поле
        'error_message': ''
    }
    
    # Определяем URL в начале функции
    product_url = f"https://autopiter.ru/goods/{quote(str(article))}"
    
    try:
        supplier_codes = SUPPLIER_CODES['autopiter']
        our_prices = []  # Будем собирать все наши цены
        our_data = []  # Будем собирать данные наших поставщиков (цена + количество)
        competitor_prices = []  # Будем собирать все цены конкурентов
        competitor_data = []  # Будем собирать данные конкурентов (цена + количество)
        
        # Значительно увеличенная задержка перед запросом, чтобы избежать rate limit
        time.sleep(random.uniform(5.0, 10.0))
        
        # Получаем прокси для запросов (если доступны)
        proxy_dict = get_next_proxy()
        proxy_str = None
        if proxy_dict:
            proxy_url = proxy_dict.get('http', '')
            if proxy_url.startswith('http://'):
                proxy_str = proxy_url[7:]  # Убираем 'http://'
                print(f"[DEBUG] Используется прокси: {proxy_str}")
        else:
            print(f"[DEBUG] Прокси не найдены, работаем без прокси")
        
        resp = make_request(product_url, proxy=proxy_str, timeout=TIMEOUT, max_retries=MAX_HTTP_RETRIES)
        if resp and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # Ищем ссылку на карточку товара
            card_link = soup.find('a', href=re.compile(r'/goods/.*/id\d+'))
            if card_link:
                card_url = 'https://autopiter.ru' + card_link['href']
                # Парсим карточку товара с тем же прокси
                card_resp = make_request(card_url, proxy=proxy_str, timeout=TIMEOUT, max_retries=MAX_HTTP_RETRIES)
                if card_resp and card_resp.status_code == 200:
                    card_soup = BeautifulSoup(card_resp.text, 'html.parser')
                    
                    print(f"[DEBUG] Карточка загружена: {card_url}")
                    
                    # Сохраняем HTML для отладки
                    with open('/tmp/autopiter_debug.html', 'w', encoding='utf-8') as f:
                        f.write(card_resp.text)
                    print(f"[DEBUG] HTML сохранен в /tmp/autopiter_debug.html")
                    
                    # Ищем минимальную цену конкурента из блока SelectedOffer
                    min_price_selectors = [
                        '.SelectedOffer__price___Xzg0ZD',
                        '.SelectedOffer__price___Xzg0ZD span',
                        'div.SelectedOffer__price___Xzg0ZD',
                        'div[class*="SelectedOffer__price"]',
                        'span[class*="SelectedOffer__price"]',
                        '.AppraiseBestItems__root___ZmRhZj .SelectedOffer__price___Xzg0ZD',
                        '.AppraiseCard__wrapper___ZmZjYm .SelectedOffer__price___Xzg0ZD',
                        '.AppraiseBestItems__root___ZmRhZj div[class*="SelectedOffer__price"]',
                        '.AppraiseCard__wrapper___ZmZjYm div[class*="SelectedOffer__price"]'
                    ]
                    
                    for selector in min_price_selectors:
                        min_price_el = card_soup.select_one(selector)
                        if min_price_el:
                            min_price_text = min_price_el.get_text(strip=True)
                            min_match = re.search(r'(\d[\d\s]*)', min_price_text.replace('\xa0', ' '))
                            if min_match:
                                competitor_min = float(min_match.group(1).replace(' ', ''))
                                print(f"[DEBUG] Найдена минимальная цена конкурента: {competitor_min}")
                                result['min_competitor_price'] = competitor_min
                                break
                    
                    # Ищем раздел "Запрошенный номер" и таблицу с предложениями только в этом разделе
                    requested_section = None
                    tables = []
                    
                    # Ищем заголовок "Запрошенный номер" (он же "Запрошенный товар")
                    section_titles = card_soup.find_all('div', class_=re.compile(r'.*NonRetailAppraiseTable__sectionTitle.*'))
                    for title_div in section_titles:
                        title_text = title_div.get_text(strip=True)
                        print(f"[DEBUG] Проверяем раздел: '{title_text}'")
                        if 'Запрошенный' in title_text:
                            print(f"[DEBUG] Найден раздел 'Запрошенный': {title_text}")
                            requested_section = title_div
                            break
                    
                    if requested_section:
                        # Ищем таблицу ТОЛЬКО в разделе "Запрошенный"
                        # Поднимаемся до контейнера таблицы
                        section_container = requested_section.find_parent('div', class_=re.compile(r'.*NonRetailAppraiseTable.*'))
                        
                        if section_container:
                            # Ищем следующую таблицу после заголовка внутри этого контейнера
                            print(f"[DEBUG] Ищем таблицу в контейнере")
                            
                            # Ищем первого sibling после заголовка, который является таблицей
                            current = requested_section
                            for _ in range(20):  # Максимум 20 итераций
                                current = current.next_sibling if hasattr(current, 'next_sibling') else None
                                if current is None:
                                    break
                                
                                # Проверяем, что мы все еще в пределах нашего контейнера
                                if section_container != current.find_parent():
                                    break
                                
                                # Проверяем, не встретили ли следующий заголовок
                                if current.name == 'div' and 'sectionTitle' in str(current.get('class', [])):
                                    print(f"[DEBUG] Достигнут следующий раздел, прекращаем поиск")
                                    break
                                
                                # Проверяем, является ли элемент таблицей
                                if current.name == 'table':
                                    tables = [current]
                                    print(f"[DEBUG] Найдена таблица в разделе 'Запрошенный'")
                                    break
                            
                            # Если не нашли таблицу как sibling, ищем в дочерних элементах
                            if not tables:
                                print(f"[DEBUG] Ищем таблицу в дочерних элементах раздела")
                                for child in section_container.children:
                                    if hasattr(child, 'name') and child.name == 'table':
                                        tables = [child]
                                        print(f"[DEBUG] Найдена таблица в дочерних элементах")
                                        break
                        else:
                            print(f"[DEBUG] Контейнер раздела не найден")
                        
                        # Если все еще не нашли, берем первую таблицу после заголовка
                        if not tables:
                            print(f"[DEBUG] Fallback: используем find_next для поиска таблицы")
                            found_tables = requested_section.find_all_next('table', limit=2)
                            if found_tables:
                                # Берем первую таблицу, которая не является дочерним элементом следующего раздела
                                for table in found_tables:
                                    # Проверяем, нет ли перед таблицей следующего заголовка
                                    prev_elements = []
                                    current = table.previous_sibling
                                    for _ in range(10):
                                        if current is None:
                                            break
                                        if hasattr(current, 'name'):
                                            prev_elements.append(current)
                                            if current.name == 'div' and 'sectionTitle' in str(current.get('class', [])):
                                                # Нашли заголовок раздела после нашей таблицы, это не наша таблица
                                                print(f"[DEBUG] Пропускаем таблицу, она не в разделе 'Запрошенный'")
                                                break
                                        current = current.previous_sibling
                                    
                                    # Если не нашли следующий заголовок, это наша таблица
                                    if not any('sectionTitle' in str(elem.get('class', [])) for elem in prev_elements if hasattr(elem, 'get')):
                                        tables = [table]
                                        print(f"[DEBUG] Найдена таблица после fallback")
                                        break
                                    
                                    if tables:
                                        break
                            
                            if not tables and found_tables:
                                # Если все fallback не сработали, берем первую таблицу
                                tables = [found_tables[0]]
                                print(f"[DEBUG] Fallback: берем первую найденную таблицу")
                    else:
                        print(f"[DEBUG] Раздел 'Запрошенный' не найден, парсим все таблицы")
                        tables = card_soup.find_all('table')
                    
                    print(f"[DEBUG] Найдено таблиц: {len(tables)}")
                    
                    for table_idx, table in enumerate(tables):
                        print(f"[DEBUG] Анализируем таблицу {table_idx + 1}")
                        
                        # Ищем все строки в таблице
                        rows = table.find_all('tr')
                        print(f"[DEBUG] В таблице {table_idx + 1} найдено строк: {len(rows)}")
                        
                        for row_idx, row in enumerate(rows):
                            try:
                                # Пропускаем заголовки
                                if row_idx == 0:
                                    continue
                                    
                                # Ищем все ячейки в строке
                                cells = row.find_all(['td', 'th'])
                                if len(cells) < 5:  # Минимум 5 колонок должно быть
                                    continue
                                
                                # Логируем содержимое строки для отладки
                                row_text = ' | '.join([cell.get_text(strip=True) for cell in cells])
                                print(f"[DEBUG] Строка {row_idx}: {row_text}")
                                
                                # Парсим данные из строки согласно структуре таблицы
                                # Регион | Поставщик | Производитель | Номер | Наименование | Наличие | Доставка | Цена | Заказ
                                if len(cells) >= 8:
                                    supplier_text = cells[1].get_text(strip=True)  # Колонка "Поставщик"
                                    
                                    # Ищем цену в правильной ячейке - пробуем несколько вариантов
                                    price_val = None
                                    
                                    print(f"[DEBUG] Ищем цену в строке {row_idx}, количество ячеек: {len(cells)}")
                                    
                                    # Сначала пробуем найти цену в правильной колонке (7-я колонка обычно содержит цену)
                                    if len(cells) > 7:
                                        price_cell = cells[7]
                                        price_text = price_cell.get_text(strip=True)
                                        print(f"[DEBUG] Проверяем колонку 7 (цена): '{price_text}'")
                                        
                                        # Ищем цену в div с классом NonRetailAppraiseTR__priceWrapper
                                        price_divs = price_cell.find_all(['div', 'span'], class_=re.compile(r'.*NonRetailAppraiseTR__priceWrapper.*'))
                                        print(f"[DEBUG] Найдено {len(price_divs)} элементов с классом NonRetailAppraiseTR__priceWrapper")
                                        
                                        # Логируем все найденные элементы для диагностики
                                        for idx, div in enumerate(price_divs):
                                            print(f"[DEBUG] Элемент {idx}: класс='{div.get('class')}', текст='{div.get_text(strip=True)}'")
                                        
                                        for price_div in price_divs:
                                            # Ищем span внутри div
                                            price_span = price_div.find('span')
                                            if price_span:
                                                div_text = price_span.get_text(strip=True)
                                                print(f"[DEBUG] Найден span с ценой: '{div_text}'")
                                            else:
                                                div_text = price_div.get_text(strip=True)
                                                print(f"[DEBUG] Найден div с ценой: '{div_text}'")
                                            
                                            # Улучшенное извлечение цены - ищем точное совпадение с рублями
                                            price_match = re.search(r'(\d[\d\s]*)\s*₽', div_text.replace('\xa0', ' '))
                                            if price_match:
                                                price_val = float(price_match.group(1).replace(' ', ''))
                                                print(f"[DEBUG] Найдена цена {price_val} в div NonRetailAppraiseTR__priceWrapper")
                                                break
                                            else:
                                                # Если не нашли с рублями, пробуем без них
                                                price_match = re.search(r'(\d[\d\s]*)', div_text.replace('\xa0', ' '))
                                                if price_match:
                                                    price_val = float(price_match.group(1).replace(' ', ''))
                                                    print(f"[DEBUG] Найдена цена {price_val} в div NonRetailAppraiseTR__priceWrapper (без ₽)")
                                                    break
                                        
                                        # Если не нашли в div, пробуем найти все span элементы в ячейке
                                        if price_val is None:
                                            all_spans = price_cell.find_all('span')
                                            print(f"[DEBUG] Найдено {len(all_spans)} span элементов в ячейке цены")
                                            for span in all_spans:
                                                span_text = span.get_text(strip=True)
                                                if '₽' in span_text:
                                                    print(f"[DEBUG] Найден span с рублями: '{span_text}'")
                                                    price_match = re.search(r'(\d[\d\s]*)\s*₽', span_text.replace('\xa0', ' '))
                                                    if price_match:
                                                        price_val = float(price_match.group(1).replace(' ', ''))
                                                        print(f"[DEBUG] Найдена цена {price_val} в span с рублями")
                                                        break
                                        
                                        # Если все еще не нашли, ищем в тексте ячейки напрямую
                                        if price_val is None:
                                            # Ищем все числа с рублями в тексте ячейки
                                            price_matches = re.findall(r'(\d[\d\s]*)\s*₽', price_text.replace('\xa0', ' '))
                                            if price_matches:
                                                # Берем первое найденное число
                                                price_val = float(price_matches[0].replace(' ', ''))
                                                print(f"[DEBUG] Найдена цена {price_val} в тексте ячейки (прямой поиск)")
                                            else:
                                                # Если не нашли с рублями, ищем просто числа
                                                number_matches = re.findall(r'(\d[\d\s]+)', price_text.replace('\xa0', ' '))
                                                if number_matches:
                                                    # Берем самое большое число (скорее всего это цена)
                                                    numbers = [float(match.replace(' ', '')) for match in number_matches]
                                                    if numbers:
                                                        price_val = max(numbers)
                                                        print(f"[DEBUG] Найдена цена {price_val} в тексте ячейки (поиск чисел)")
                                        
                                        # Если не нашли в div, ищем в тексте ячейки
                                        if price_val is None:
                                            price_match = re.search(r'(\d[\d\s]*)\s*₽', price_text)
                                            if price_match:
                                                price_val = float(price_match.group(1).replace(' ', ''))
                                                print(f"[DEBUG] Найдена цена {price_val} в тексте ячейки")
                                    
                                    # Если не нашли, пробуем другие ячейки с ценой
                                    if price_val is None:
                                        for cell_idx, cell in enumerate(cells):
                                            cell_class = str(cell.get('class', []))
                                            cell_text = cell.get_text(strip=True)
                                            
                                            # Пропускаем ячейки с кодом поставщика и региона
                                            if 'supplier' in cell_class.lower() or 'region' in cell_class.lower():
                                                continue
                                            
                                            # Ищем ячейки с классом price
                                            if 'price' in cell_class.lower():
                                                price_match = re.search(r'(\d[\d\s]*)\s*₽', cell_text)
                                                if price_match:
                                                    price_val = float(price_match.group(1).replace(' ', ''))
                                                    print(f"[DEBUG] Найдена цена {price_val} в ячейке с классом price")
                                                    break
                                    
                                    print(f"[DEBUG] Итоговая найденная цена: {price_val}")
                                    
                                    # Извлекаем количество товара из ячейки с количеством
                                    quantity_val = None
                                    if len(cells) >= 6:  # Колонка с количеством обычно 5-я (индекс 5)
                                        quantity_cell = cells[5]  # Колонка "Наличие"
                                        quantity_text = quantity_cell.get_text(strip=True)
                                        print(f"[DEBUG] Проверяем колонку 5 (количество): '{quantity_text}'")
                                        
                                        # Ищем количество в div с классом NonRetailAppraiseTR__quantity
                                        quantity_divs = quantity_cell.find_all(['div', 'span'], class_=re.compile(r'.*NonRetailAppraiseTR__quantity.*'))
                                        for quantity_div in quantity_divs:
                                            # Ищем span внутри div
                                            quantity_span = quantity_div.find('span')
                                            if quantity_span:
                                                div_text = quantity_span.get_text(strip=True)
                                                print(f"[DEBUG] Найден span с количеством: '{div_text}'")
                                            else:
                                                div_text = quantity_div.get_text(strip=True)
                                                print(f"[DEBUG] Найден div с количеством: '{div_text}'")
                                            
                                            # Извлекаем число из текста типа "32 шт" или ">20 шт"
                                            quantity_match = re.search(r'(\d+)', div_text)
                                            if quantity_match:
                                                quantity_val = int(quantity_match.group(1))
                                                print(f"[DEBUG] Найдено количество {quantity_val} в div NonRetailAppraiseTR__quantity")
                                                break
                                        
                                        # Если не нашли в div, ищем в тексте ячейки
                                        if quantity_val is None:
                                            quantity_match = re.search(r'(\d+)', quantity_text)
                                            if quantity_match:
                                                quantity_val = int(quantity_match.group(1))
                                                print(f"[DEBUG] Найдено количество {quantity_val} в тексте ячейки")
                                    
                                    print(f"[DEBUG] Итоговое найденное количество: {quantity_val}")
                                    
                                    if price_val is not None:
                                        # Извлекаем цифры из кода поставщика
                                        sup_digits = re.sub(r'\D+', '', supplier_text)
                                        
                                        print(f"[DEBUG] Найден поставщик '{supplier_text}' ({sup_digits}) с ценой {price_val}")
                                        print(f"[DEBUG] Наши коды поставщиков: {supplier_codes}")
                                        
                                        # Более точная проверка кода поставщика
                                        is_our_supplier = False
                                        if sup_digits:
                                            # Проверяем точное совпадение
                                            if sup_digits in supplier_codes:
                                                is_our_supplier = True
                                                print(f"[DEBUG] ✅ ТОЧНОЕ СОВПАДЕНИЕ НАЙДЕНО: {sup_digits} в наших кодах!")
                                            else:
                                                # Проверяем, содержит ли код поставщика наши коды как подстроку
                                                for our_code in supplier_codes:
                                                    if our_code in sup_digits:
                                                        is_our_supplier = True
                                                        print(f"[DEBUG] ✅ ПОДСТРОКА НАЙДЕНА: {our_code} в {sup_digits}")
                                                        break
                                                if not is_our_supplier:
                                                    print(f"[DEBUG] ❌ Поставщик {sup_digits} НЕ найден в наших кодах {supplier_codes}")
                                        else:
                                            print(f"[DEBUG] ❌ Не удалось извлечь цифры из кода поставщика: '{supplier_text}'")
                                        
                                        print(f"[DEBUG] Это наш поставщик: {is_our_supplier}")
                                        
                                        if is_our_supplier:
                                            our_prices.append(price_val)
                                            our_data.append({
                                                'price': price_val,
                                                'quantity': quantity_val,
                                                'supplier': sup_digits
                                            })
                                            print(f"[DEBUG] Добавлена наша цена {price_val} для поставщика {sup_digits}, количество: {quantity_val}")
                                        else:
                                            competitor_prices.append(price_val)
                                            competitor_data.append({
                                                'price': price_val,
                                                'quantity': quantity_val,
                                                'supplier': sup_digits
                                            })
                                            print(f"[DEBUG] Добавлена цена конкурента {price_val} для поставщика {sup_digits}, количество: {quantity_val}")
                                            
                            except Exception as e:
                                print(f"[DEBUG] Ошибка парсинга строки {row_idx}: {str(e)}")
                                continue
                    
                    # Обрабатываем найденные цены
                    print(f"[DEBUG] === ИТОГОВАЯ СТАТИСТИКА ===")
                    print(f"[DEBUG] Найдено наших поставщиков: {len(our_prices)}")
                    print(f"[DEBUG] Найдено конкурентов: {len(competitor_prices)}")
                    print(f"[DEBUG] Наши коды: {supplier_codes}")
                    print(f"[DEBUG] =========================")
                    
                    if our_prices:
                        # Берем минимальную цену среди наших поставщиков
                        result['marketplace_price'] = min(our_prices)
                        result['is_found'] = True
                        print(f"[DEBUG] Найдены наши цены: {our_prices}, минимальная: {result['marketplace_price']}")
                        
                        # Находим предложение нашего поставщика с минимальной ценой и устанавливаем количество
                        for data in our_data:
                            if data['price'] == result['marketplace_price']:
                                result['quantity_in_stock'] = data['quantity']
                                print(f"[DEBUG] Установлено количество от нашего поставщика: {result['quantity_in_stock']}")
                                break
                        
                        # Если не нашли количество для минимальной цены, берем первое доступное
                        if result['quantity_in_stock'] is None and our_data:
                            result['quantity_in_stock'] = our_data[0]['quantity']
                            print(f"[DEBUG] Установлено количество от первого нашего поставщика: {result['quantity_in_stock']}")
                    else:
                        print(f"[DEBUG] ❌ НАШИ ЦЕНЫ НЕ НАЙДЕНЫ! Наши коды поставщиков: {supplier_codes}")
                        print(f"[DEBUG] Возможно, нужно обновить коды поставщиков или наши поставщики не представлены на этой странице")
                    
                    # Если нашли минимальную цену конкурента из таблицы, используем её
                    if competitor_data:
                        # Находим предложение конкурента с минимальной ценой
                        competitor_offers = [d for d in competitor_data if d['supplier'] not in supplier_codes]
                        if competitor_offers:
                            min_competitor_offer = min(competitor_offers, key=lambda x: x['price'])
                            result['min_competitor_price'] = min_competitor_offer['price']
                            # Сохраняем количество для минимальной цены конкурента
                            result['competitor_quantity'] = min_competitor_offer['quantity']
                            print(f"[DEBUG] Установлена минимальная цена конкурента из таблицы: {result['min_competitor_price']}, количество: {result['competitor_quantity']}")
                        else:
                            # Если не нашли конкурентов, но есть данные, берем минимальную цену из всех данных
                            if competitor_data:
                                min_competitor_offer = min(competitor_data, key=lambda x: x['price'])
                                result['min_competitor_price'] = min_competitor_offer['price']
                                result['competitor_quantity'] = min_competitor_offer['quantity']
                                print(f"[DEBUG] Установлена минимальная цена из всех данных: {result['min_competitor_price']}, количество: {result['competitor_quantity']}")
                            else:
                                print(f"[DEBUG] Нет данных конкурентов для установки минимальной цены")
                    
                    if result['min_competitor_price'] is None:
                        print(f"[DEBUG] Минимальная цена конкурента не найдена")
                    
    except Exception as e:
        result['error_message'] = f'HTTP parsing failed: {str(e)}'
        print(f"[DEBUG] Ошибка HTTP парсинга: {str(e)}")
    
    # Если HTTP не нашел нужные элементы, пробуем Selenium с прокси
    if not result['is_found']:
        # Дополнительная задержка перед Selenium для снижения нагрузки
        time.sleep(random.uniform(3.0, 6.0))
        driver = None
        try:
            options = Options()
            options.add_argument('--headless=new')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--disable-web-security')
            options.add_argument('--disable-features=VizDisplayCompositor')
            options.add_argument('--remote-debugging-port=9222')
            
            # Добавляем прокси для Selenium, если доступен
            proxy_dict = get_next_proxy()
            if proxy_dict:
                proxy_url = proxy_dict.get('http', '')
                if proxy_url.startswith('http://'):
                    proxy_host = proxy_url[7:]  # Убираем 'http://'
                    options.add_argument(f'--proxy-server=http://{proxy_host}')
            
            driver = webdriver.Chrome(options=options)
            driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
            driver.implicitly_wait(3)
            driver.get(product_url)
            
            # Переходим в первую карточку товара из списка
            try:
                link_el = WebDriverWait(driver, 6).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href*="/goods/"][href*="/id"]'))
                )
                href = link_el.get_attribute('href')
                if href:
                    driver.get(href)
            except Exception:
                pass
            
            # Ждем загрузки таблицы - пробуем разные селекторы
            table = None
            table_selectors = [
                "table",
                "div[class*='Table__table']",
                "div[class*='AppraiseTable']",
                "div[class*='NonRetailAppraiseTable']",
                ".AppraiseTable",
                ".NonRetailAppraiseTable"
            ]
            
            for selector in table_selectors:
                try:
                    table = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    print(f"[DEBUG] Selenium: таблица найдена по селектору '{selector}'")
                    break
                except Exception as e:
                    print(f"[DEBUG] Selenium: не найдена таблица по селектору '{selector}': {str(e)}")
                    continue
            
            if not table:
                print(f"[DEBUG] Selenium: таблица не найдена ни по одному селектору")
                raise Exception("Таблица не найдена")
            
            # Ищем минимальную цену конкурента
            if result['min_competitor_price'] is None:
                min_price_selectors = [
                    '.SelectedOffer__price___Xzg0ZD',
                    '.SelectedOffer__price___Xzg0ZD span',
                    'div.SelectedOffer__price___Xzg0ZD',
                    'div[class*="SelectedOffer__price"]',
                    'span[class*="SelectedOffer__price"]',
                    '.AppraiseBestItems__root___ZmRhZj .SelectedOffer__price___Xzg0ZD',
                    '.AppraiseCard__wrapper___ZmZjYm .SelectedOffer__price___Xzg0ZD',
                    '.AppraiseBestItems__root___ZmRhZj div[class*="SelectedOffer__price"]',
                    '.AppraiseCard__wrapper___ZmZjYm div[class*="SelectedOffer__price"]'
                ]
                
                for selector in min_price_selectors:
                    try:
                        min_price_el = driver.find_element(By.CSS_SELECTOR, selector)
                        min_price_text = min_price_el.text.strip()
                        min_match = re.search(r'(\d[\d\s]*)', min_price_text.replace('\xa0', ' '))
                        if min_match:
                            result['min_competitor_price'] = float(min_match.group(1).replace(' ', ''))
                            print(f"[DEBUG] Selenium: найдена минимальная цена конкурента: {result['min_competitor_price']}")
                            break
                    except Exception as e:
                        print(f"[DEBUG] Selenium: не удалось найти минимальную цену по селектору {selector}: {str(e)}")
                        continue
            
            # Парсим таблицу с предложениями
            rows = table.find_elements(By.TAG_NAME, "tr")
            print(f"[DEBUG] Selenium: найдено строк в таблице: {len(rows)}")
            
            for row_idx, row in enumerate(rows):
                try:
                    if row_idx == 0:  # Пропускаем заголовок
                        continue
                        
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) < 8:
                        continue
                        
                    # Получаем текст из ячеек
                    supplier_text = cells[1].text.strip()  # Поставщик
                    
                    # Ищем цену в правильной ячейке - пробуем несколько вариантов
                    price_val = None
                    
                    print(f"[DEBUG] Selenium: ищем цену в строке {row_idx}, количество ячеек: {len(cells)}")
                    
                    # Сначала пробуем найти цену в правильной колонке (7-я колонка обычно содержит цену)
                    if len(cells) > 7:
                        price_cell = cells[7]
                        price_text = price_cell.text.strip()
                        print(f"[DEBUG] Selenium: проверяем колонку 7 (цена): '{price_text}'")
                        
                        # Ищем цену в div с классом NonRetailAppraiseTR__priceWrapper
                        try:
                            price_divs = price_cell.find_elements(By.CSS_SELECTOR, 'div[class*="NonRetailAppraiseTR__priceWrapper"], span[class*="NonRetailAppraiseTR__priceWrapper"]')
                            print(f"[DEBUG] Selenium: найдено {len(price_divs)} элементов с классом NonRetailAppraiseTR__priceWrapper")
                            
                            # Логируем все найденные элементы для диагностики
                            for idx, div in enumerate(price_divs):
                                try:
                                    class_attr = div.get_attribute('class')
                                    text_content = div.text.strip()
                                    print(f"[DEBUG] Selenium: элемент {idx}: класс='{class_attr}', текст='{text_content}'")
                                except Exception as e:
                                    print(f"[DEBUG] Selenium: ошибка получения атрибутов элемента {idx}: {str(e)}")
                            
                            for price_div in price_divs:
                                div_text = price_div.text.strip()
                                print(f"[DEBUG] Selenium: проверяем элемент: '{div_text}'")
                                # Улучшенное извлечение цены - ищем точное совпадение с рублями
                                price_match = re.search(r'(\d[\d\s]*)\s*₽', div_text.replace('\xa0', ' '))
                                if price_match:
                                    price_val = float(price_match.group(1).replace(' ', ''))
                                    print(f"[DEBUG] Selenium: найдена цена {price_val} в div NonRetailAppraiseTR__priceWrapper")
                                    break
                                else:
                                    # Если не нашли с рублями, пробуем без них
                                    price_match = re.search(r'(\d[\d\s]*)', div_text.replace('\xa0', ' '))
                                    if price_match:
                                        price_val = float(price_match.group(1).replace(' ', ''))
                                        print(f"[DEBUG] Selenium: найдена цена {price_val} в div NonRetailAppraiseTR__priceWrapper (без ₽)")
                                        break
                            
                            # Если не нашли в div, пробуем найти все span элементы в ячейке
                            if price_val is None:
                                all_spans = price_cell.find_elements(By.TAG_NAME, 'span')
                                print(f"[DEBUG] Selenium: найдено {len(all_spans)} span элементов в ячейке цены")
                                for span in all_spans:
                                    span_text = span.text.strip()
                                    if '₽' in span_text:
                                        print(f"[DEBUG] Selenium: найден span с рублями: '{span_text}'")
                                        price_match = re.search(r'(\d[\d\s]*)\s*₽', span_text.replace('\xa0', ' '))
                                        if price_match:
                                            price_val = float(price_match.group(1).replace(' ', ''))
                                            print(f"[DEBUG] Selenium: найдена цена {price_val} в span с рублями")
                                            break
                            
                            # Если все еще не нашли, ищем в тексте ячейки напрямую
                            if price_val is None:
                                # Ищем все числа с рублями в тексте ячейки
                                price_matches = re.findall(r'(\d[\d\s]*)\s*₽', price_text.replace('\xa0', ' '))
                                if price_matches:
                                    # Берем первое найденное число
                                    price_val = float(price_matches[0].replace(' ', ''))
                                    print(f"[DEBUG] Selenium: найдена цена {price_val} в тексте ячейки (прямой поиск)")
                                else:
                                    # Если не нашли с рублями, ищем просто числа
                                    number_matches = re.findall(r'(\d[\d\s]+)', price_text.replace('\xa0', ' '))
                                    if number_matches:
                                        # Берем самое большое число (скорее всего это цена)
                                        numbers = [float(match.replace(' ', '')) for match in number_matches]
                                        if numbers:
                                            price_val = max(numbers)
                                            print(f"[DEBUG] Selenium: найдена цена {price_val} в тексте ячейки (поиск чисел)")
                        except Exception as e:
                            print(f"[DEBUG] Selenium: ошибка поиска цены в div: {str(e)}")
                            pass
                        
                        # Если не нашли в div, ищем в тексте ячейки
                        if price_val is None:
                            price_match = re.search(r'(\d[\d\s]*)\s*₽', price_text)
                            if price_match:
                                price_val = float(price_match.group(1).replace(' ', ''))
                                print(f"[DEBUG] Selenium: найдена цена {price_val} в тексте ячейки")
                    
                    # Если не нашли, пробуем другие ячейки с ценой
                    if price_val is None:
                        for cell_idx, cell in enumerate(cells):
                            cell_class = cell.get_attribute('class') or ''
                            cell_text = cell.text.strip()
                            
                            # Пропускаем ячейки с кодом поставщика и региона
                            if 'supplier' in cell_class.lower() or 'region' in cell_class.lower():
                                continue
                            
                            # Ищем ячейки с классом price
                            if 'price' in cell_class.lower():
                                price_match = re.search(r'(\d[\d\s]*)\s*₽', cell_text)
                                if price_match:
                                    price_val = float(price_match.group(1).replace(' ', ''))
                                    print(f"[DEBUG] Selenium: найдена цена {price_val} в ячейке с классом price")
                                    break
                    
                    print(f"[DEBUG] Selenium: итоговая найденная цена: {price_val}")
                    
                    # Извлекаем количество товара из ячейки с количеством
                    quantity_val = None
                    if len(cells) >= 6:  # Колонка с количеством обычно 5-я (индекс 5)
                        quantity_cell = cells[5]  # Колонка "Наличие"
                        quantity_text = quantity_cell.text.strip()
                        print(f"[DEBUG] Selenium: проверяем колонку 5 (количество): '{quantity_text}'")
                        
                        # Ищем количество в div с классом NonRetailAppraiseTR__quantity
                        try:
                            quantity_divs = quantity_cell.find_elements(By.CSS_SELECTOR, 'div[class*="NonRetailAppraiseTR__quantity"], span[class*="NonRetailAppraiseTR__quantity"]')
                            for quantity_div in quantity_divs:
                                div_text = quantity_div.text.strip()
                                print(f"[DEBUG] Selenium: найден div с количеством: '{div_text}'")
                                
                                # Извлекаем число из текста типа "32 шт" или ">20 шт"
                                quantity_match = re.search(r'(\d+)', div_text)
                                if quantity_match:
                                    quantity_val = int(quantity_match.group(1))
                                    print(f"[DEBUG] Selenium: найдено количество {quantity_val} в div NonRetailAppraiseTR__quantity")
                                    break
                            
                            # Если не нашли в div, ищем в тексте ячейки
                            if quantity_val is None:
                                quantity_match = re.search(r'(\d+)', quantity_text)
                                if quantity_match:
                                    quantity_val = int(quantity_match.group(1))
                                    print(f"[DEBUG] Selenium: найдено количество {quantity_val} в тексте ячейки")
                        except Exception as e:
                            print(f"[DEBUG] Selenium: ошибка поиска количества: {str(e)}")
                    
                    print(f"[DEBUG] Selenium: итоговое найденное количество: {quantity_val}")
                    
                    if price_val is not None:
                        # Извлекаем цифры из кода поставщика
                        sup_digits = re.sub(r'\D+', '', supplier_text)
                        
                        print(f"[DEBUG] Selenium: найден поставщик '{supplier_text}' ({sup_digits}) с ценой {price_val}")
                        print(f"[DEBUG] Selenium: наши коды поставщиков: {supplier_codes}")
                        
                        # Более точная проверка кода поставщика
                        is_our_supplier = False
                        if sup_digits:
                            # Проверяем точное совпадение
                            if sup_digits in supplier_codes:
                                is_our_supplier = True
                                print(f"[DEBUG] Selenium: ✅ ТОЧНОЕ СОВПАДЕНИЕ НАЙДЕНО: {sup_digits} в наших кодах!")
                            else:
                                # Проверяем, содержит ли код поставщика наши коды как подстроку
                                for our_code in supplier_codes:
                                    if our_code in sup_digits:
                                        is_our_supplier = True
                                        print(f"[DEBUG] Selenium: ✅ ПОДСТРОКА НАЙДЕНА: {our_code} в {sup_digits}")
                                        break
                                if not is_our_supplier:
                                    print(f"[DEBUG] Selenium: ❌ Поставщик {sup_digits} НЕ найден в наших кодах {supplier_codes}")
                        else:
                            print(f"[DEBUG] Selenium: ❌ Не удалось извлечь цифры из кода поставщика: '{supplier_text}'")
                        
                        print(f"[DEBUG] Selenium: это наш поставщик: {is_our_supplier}")
                        
                        if is_our_supplier:
                            our_prices.append(price_val)
                            our_data.append({
                                'price': price_val,
                                'quantity': quantity_val,
                                'supplier': sup_digits
                            })
                            print(f"[DEBUG] Selenium: добавлена наша цена {price_val} для поставщика {sup_digits}, количество: {quantity_val}")
                        else:
                            competitor_prices.append(price_val)
                            competitor_data.append({
                                'price': price_val,
                                'quantity': quantity_val,
                                'supplier': sup_digits
                            })
                            print(f"[DEBUG] Selenium: добавлена цена конкурента {price_val} для поставщика {sup_digits}, количество: {quantity_val}")
                            
                except Exception as e:
                    print(f"[DEBUG] Selenium: ошибка парсинга строки {row_idx}: {str(e)}")
                    continue
            
            # Обрабатываем найденные цены
            if our_prices:
                result['marketplace_price'] = min(our_prices)
                result['is_found'] = True
                print(f"[DEBUG] Selenium: найдены наши цены: {our_prices}, минимальная: {result['marketplace_price']}")
            
            # Если не нашли минимальную цену конкурента, используем минимальную из таблицы
            if result['min_competitor_price'] is None and competitor_prices:
                result['min_competitor_price'] = min(competitor_prices)
                print(f"[DEBUG] Selenium: установлена минимальная цена конкурента из таблицы: {result['min_competitor_price']}")
            
            # Устанавливаем количество для Selenium (если не установлено)
            if result['quantity_in_stock'] is None and our_data:
                result['quantity_in_stock'] = our_data[0]['quantity']
                print(f"[DEBUG] Selenium: установлено количество от первого нашего поставщика: {result['quantity_in_stock']}")
            
            if result['competitor_quantity'] is None and competitor_data:
                result['competitor_quantity'] = competitor_data[0]['quantity']
                print(f"[DEBUG] Selenium: установлено количество от первого конкурента: {result['competitor_quantity']}")
                
        except Exception as e:
            if not result['error_message']:
                result['error_message'] = f'Selenium failed: {str(e)}'
            print(f"[DEBUG] Selenium: ошибка поиска таблицы: {str(e)}")
        finally:
            try:
                if driver:
                    driver.quit()
            except Exception:
                pass
    
    print(f"[DEBUG] ===== КОНЕЦ check_autopiter_item =====")
    print(f"[DEBUG] Итоговый результат: наш={result['marketplace_price']}, конкурент={result['min_competitor_price']}, найдено={result['is_found']}")
    print(f"[DEBUG] Количество в наличии: {result['quantity_in_stock']}")
    print(f"[DEBUG] Количество конкурента: {result['competitor_quantity']}")
    print(f"[DEBUG] Полный результат: {result}")
    return result

def check_emex_item(supplier_code: str, manufacturer: str, article: str, competitor_brand_filter: str = None) -> Dict:
    """Проверяет наличие позиции на Emex и анализирует цены.

    Если передан код поставщика (как на странице data-vendor-id/data-vendor-code),
    пытается найти оффер конкретного поставщика и извлечь его цену.
    При отсутствии кода — использует общую минимальную цену из выдачи.
    """
    result = {
        'is_found': False,
        'marketplace_price': None,
        'min_competitor_price': None,
        'competitor_brand': None,
        'quantity_in_stock': None,
        'error_message': ''
    }

    try:
        # ------ Selenium-first flow per new Emex requirements ------
        try:
            temp_dir = tempfile.mkdtemp(prefix=f"chrome_emex_{uuid.uuid4().hex[:8]}_")
            driver = _create_chrome_driver_robust(temp_dir, proxy=None)
        except Exception:
            driver = None
            temp_dir = None

        if driver:
            try:
                products_url = f"https://emex.ru/products/{quote(article)}"
                print(f"[DEBUG] Emex Selenium: открываем {products_url}")
                driver.get(products_url)
                WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#__next')))

                # Кликаем "Все предложения"
                try:
                    all_offers_btn = WebDriverWait(driver, 6).until(
                        EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Все предложения')]"))
                    )
                    clickable = all_offers_btn
                    try:
                        clickable = all_offers_btn.find_element(By.XPATH, './ancestor::a')
                    except Exception:
                        pass
                    driver.execute_script("arguments[0].click();", clickable)
                    time.sleep(2)
                except Exception as e:
                    print(f"[DEBUG] Emex Selenium: кнопка 'Все предложения' не найдена: {str(e)}")

                # Ждем блок "Искомый товар"
                section_header = None
                try:
                    section_header = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//h2[contains(text(),'Искомый товар') or @data-testid='Offers:test:originalsTableTitle']"))
                    )
                    print("[DEBUG] Emex Selenium: секция 'Искомый товар' найдена")
                except Exception:
                    print("[DEBUG] Emex Selenium: секция 'Искомый товар' не найдена")

                offers: List[Dict] = []
                if section_header:
                    # Ищем строки предложений ТОЛЬКО внутри секции "Искомый товар"
                    price_nodes = driver.find_elements(By.CSS_SELECTOR, "[data-testid='Offers:text:priceInfo']")
                    print(f"[DEBUG] Emex Selenium: найдено ценовых узлов: {len(price_nodes)}")
                    for p in price_nodes:
                        try:
                            # фильтр: цена должна находиться внутри контейнера, где в предках есть заголовок "Искомый товар"
                            in_section = False
                            try:
                                sec_candidates = p.find_elements(By.XPATH, "ancestor::*[.//h2[contains(., 'Искомый товар') or @data-testid='Offers:test:originalsTableTitle']]")
                                in_section = len(sec_candidates) > 0
                            except Exception:
                                in_section = False
                            if not in_section:
                                continue

                            # ближайший общий контейнер строки: содержит и цену, и количество
                            try:
                                container = p.find_element(By.XPATH, "ancestor::*[.//*[@data-testid='Offers:text:priceInfo'] and .//*[@data-testid='Offers:text:quantityInfo']][1]")
                            except Exception:
                                container = p
                                for _ in range(8):
                                    try:
                                        container = container.find_element(By.XPATH, './..')
                                    except Exception:
                                        break

                            price_text = p.text.strip()
                            m = re.search(r"(\d[\d\s]*)", price_text.replace('\xa0',' '))
                            if not m:
                                continue
                            price_val = float(m.group(1).replace(' ', ''))

                            qty_val = None
                            try:
                                qty_el = container.find_element(By.CSS_SELECTOR, "[data-testid='Offers:text:quantityInfo']")
                                qty_text = qty_el.text.strip()
                                mqty = re.search(r"(\d+)", qty_text)
                                if mqty:
                                    qty_val = int(mqty.group(1))
                            except Exception:
                                # Fallback: ищем quantityInfo внутри той же секции, ближайший к текущей цене
                                try:
                                    base_sec = p.find_element(By.XPATH, "ancestor::*[.//h2[contains(., 'Искомый товар') or @data-testid='Offers:test:originalsTableTitle']][1]")
                                    qty_candidates = base_sec.find_elements(By.CSS_SELECTOR, "[data-testid='Offers:text:quantityInfo']")
                                    if qty_candidates:
                                        # берем первый как наиболее вероятный (в списке DOM-узлов секции порядок сохранен)
                                        qty_text = qty_candidates[0].text.strip()
                                        mqty = re.search(r"(\d+)", qty_text)
                                        if mqty:
                                            qty_val = int(mqty.group(1))
                                except Exception:
                                    qty_val = None

                            if 50 <= price_val <= 200000:
                                offers.append({'price': price_val, 'quantity': qty_val})
                                print(f"[DEBUG] Emex Selenium: оффер цена={price_val}, qty={qty_val}")
                        except Exception as e:
                            print(f"[DEBUG] Emex Selenium: ошибка разбора оффера: {str(e)}")
                            continue

                    # Сбор кодов поставщиков по звездочкам ТОЛЬКО в пределах секции
                    supplier_codes_found: Set[str] = set()
                    try:
                        # Ищем все звезды и фильтруем по принадлежности той же секции "Искомый товар"
                        stars_all = driver.find_elements(By.CSS_SELECTOR, "[data-testid='Offers:text:ratingInfo']")
                        stars = []
                        for st in stars_all:
                            try:
                                sec_candidates = st.find_elements(By.XPATH, "ancestor::*[.//h2[contains(., 'Искомый товар') or @data-testid='Offers:test:originalsTableTitle']]")
                                if len(sec_candidates) > 0:
                                    stars.append(st)
                            except Exception:
                                continue
                        print(f"[DEBUG] Emex Selenium: звезд найдено в секции: {len(stars)}")
                        for idx, star in enumerate(stars[:40]):
                            try:
                                driver.execute_script("arguments[0].click();", star)
                                code_el = WebDriverWait(driver, 5).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.sc-9641247d-8.hWwuoa'))
                                )
                                code_text = code_el.text.strip()
                                mm = re.match(r"([A-Z0-9]+)", code_text.replace('Ё','E').upper())
                                if mm:
                                    supplier_codes_found.add(mm.group(1))
                                driver.execute_script('document.body.click();')
                                time.sleep(0.15)
                            except Exception:
                                continue
                    except Exception:
                        pass

                    # Определяем, является ли какое-то предложение нашим
                    our_codes_upper = set(c.upper() for c in SUPPLIER_CODES.get('emex', []))
                    if supplier_code:
                        our_codes_upper.add(str(supplier_code).upper())

                    is_our_present = bool(our_codes_upper.intersection(supplier_codes_found))

                    if offers:
                        offers.sort(key=lambda x: x['price'])
                        best = offers[0]
                        result['marketplace_price'] = best['price']
                        result['quantity_in_stock'] = best['quantity']
                        result['is_found'] = True
                        print(f"[DEBUG] Emex Selenium: min price={best['price']}, qty={best['quantity']}")

                    # Если это наш поставщик найден — оставим как есть; иначе можно трактовать как конкурентную цену
                    if not is_our_present and offers:
                        result['min_competitor_price'] = result['marketplace_price']
                        result['competitor_quantity'] = result['quantity_in_stock']

                # Завершаем работу драйвера
                try:
                    driver.quit()
                except Exception:
                    pass
                if temp_dir and os.path.exists(temp_dir):
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except Exception:
                        pass

                # Если Selenium успешно нашел цену — выходим без HTTP
                if result['is_found'] and result['marketplace_price'] is not None:
                    return result

            except Exception as e:
                print(f"[DEBUG] Emex Selenium: ошибка сценария: {str(e)}")
                try:
                    driver.quit()
                except Exception:
                    pass
                if temp_dir and os.path.exists(temp_dir):
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except Exception:
                        pass

        # ------ HTTP fallback (старый путь) ------
        # Проверяем наличие брендов через основной Emex-парсер (помогает понять релевантность)
        man_ok = False
        try:
            brands = get_brands_by_artikul_emex(article) or []
            man_norm = _norm_brand(manufacturer)
            for b in brands:
                if _norm_brand(b) == man_norm:
                    man_ok = True
                    result['competitor_brand'] = b
                    break
        except Exception:
            pass

        # Формируем URL Emex на основе detailNum (стабильный вариант страницы предложений)
        # Код поставщика используем только для фильтра в DOM, не добавляем в URL
        url = f"https://emex.ru/search?detailNum={quote(article)}"

        # Делаем запрос
        response = make_request(url, timeout=TIMEOUT, max_retries=MAX_HTTP_RETRIES)
        if not response:
            result['error_message'] = "Ошибка запроса к Emex"
            return result

        soup = BeautifulSoup(response.text, 'html.parser')

        def extract_price_from_text(text: str) -> Optional[float]:
            try:
                m = re.search(r"(\d[\d\s]{2,})\s*₽", text)
                if m:
                    return float(m.group(1).replace(' ', ''))
            except Exception:
                return None
            return None

        offer_price: Optional[float] = None

        # Если известен код поставщика — ищем оффер с соответствующим data-атрибутом
        if supplier_code:
            # Ищем узлы с атрибутами data-vendor-id/data-vendor-code равными supplier_code
            vendor_nodes = soup.find_all(attrs={
                'data-vendor-id': supplier_code
            }) or soup.find_all(attrs={
                'data-vendor-code': supplier_code
            })

            for node in vendor_nodes:
                # Подымаемся к карточке оффера и ищем цену внутри блока
                container = node
                # Ограничим глубину подъема, чтобы не выйти за пределы нужного блока
                for _ in range(5):
                    if container is None:
                        break
                    # Ищем все видимые тексты с символом рубля внутри контейнера
                    price_texts = [t for t in container.stripped_strings if '₽' in t]
                    for pt in price_texts:
                        price_val = extract_price_from_text(pt)
                        if price_val:
                            offer_price = price_val
                            break
                    if offer_price is not None:
                        break
                    container = container.parent
                if offer_price is not None:
                    break

        # Фолбэк: если по коду поставщика не нашли, берём минимальную цену со всей страницы
        if offer_price is None:
            txt = soup.get_text(' ', strip=True)
            # «от N ₽» или просто «N ₽»
            m = re.search(r'от\s*(\d[\d\s]*)\s*₽', txt) or re.search(r'(\d[\d\s]{2,})\s*₽', txt)
            if m:
                try:
                    offer_price = float(m.group(1).replace(' ', ''))
                except Exception:
                    offer_price = None

        # Заполняем результат
        if offer_price is not None:
            result['marketplace_price'] = offer_price
            result['is_found'] = True
        else:
            # Товар считается найденным только если есть цена
            # Не устанавливаем is_found = True только на основе бренда
            result['is_found'] = False
            # Но сохраняем информацию о бренде, если он найден
            if man_ok:
                print(f"[DEBUG] Emex: бренд найден, но цена не найдена")

    except Exception as e:
        result['error_message'] = f"Ошибка парсинга Emex: {str(e)}"

    return result

def check_armtek_item(supplier_code: str, manufacturer: str, article: str, competitor_brand_filter: str = None) -> Dict:
    """Проверяет наличие позиции на Армтек и анализирует цены"""
    result = {
        'is_found': False,
        'marketplace_price': None,
        'min_competitor_price': None,
        'competitor_brand': None,
        'quantity_in_stock': None,
        'competitor_quantity': None,
        'error_message': ''
    }
    
    driver = None
    temp_dir = None
    
    try:
        print(f"[DEBUG] Armtek: начинаем поиск товара {manufacturer} {article}")
        
        # Создаем драйвер для парсинга
        temp_dir = tempfile.mkdtemp(prefix=f"chrome_armtek_{uuid.uuid4().hex[:8]}_")
        driver = _create_chrome_driver_robust(temp_dir, proxy=None)
        
        if not driver:
            print(f"[DEBUG] Armtek: не удалось создать драйвер")
            result['error_message'] = "Не удалось создать драйвер для парсинга"
            return result
        
        # Переходим на страницу поиска
        search_url = f"https://armtek.ru/search?text={quote(article)}"
        print(f"[DEBUG] Armtek: загружаем страницу поиска {search_url}")
        driver.get(search_url)
        
        # Ждем загрузки страницы
        time.sleep(5)
        
        # Ищем секцию "Искомый товар" и берем цены прямо со страницы поиска
        try:
            # Ищем заголовок "Искомый товар"
            target_section = driver.find_element(By.XPATH, "//p[contains(text(), 'Искомый товар')]")
            print(f"[DEBUG] Armtek: найдена секция 'Искомый товар'")
            
            # Ищем карточку товара в этой секции
            product_cards = driver.find_elements(By.CSS_SELECTOR, 'p.font__headline6')
            print(f"[DEBUG] Armtek: найдено {len(product_cards)} карточек товаров")
            
            product_found = False
            
            for card in product_cards:
                title = card.text.strip()
                print(f"[DEBUG] Armtek: проверяем карточку: '{title}'")
                if article.lower() in title.lower() and manufacturer.lower() in title.lower():
                    product_found = True
                    print(f"[DEBUG] Armtek: найдена подходящая карточка: '{title}'")
                    break
            
            if product_found:
                print(f"[DEBUG] Armtek: товар найден в секции 'Искомый товар'")
                
                # Извлекаем бренд из названия товара
                brand_text = manufacturer  # Используем переданный бренд как найденный
                print(f"[DEBUG] Armtek: используем бренд из названия товара: '{brand_text}'")
                
                # Проверяем соответствие бренда
                man_norm = _norm_brand(manufacturer)
                found_brand_norm = _norm_brand(brand_text)
                
                if man_norm == found_brand_norm:
                    result['is_found'] = True
                    result['competitor_brand'] = brand_text
                    print(f"[DEBUG] Armtek: товар найден, бренд совпадает")
                    
                    # Ищем цены прямо в секции "Искомый товар" на странице поиска
                    try:
                        # Ищем все элементы с ценами в секции "Искомый товар"
                        # Используем более точный селектор для поиска цен в этой секции
                        price_elements = driver.find_elements(By.CSS_SELECTOR, 'div.suggestion-item span.font__headline6.no-wrap')
                        print(f"[DEBUG] Armtek: найдено {len(price_elements)} элементов с ценами в секции 'Искомый товар'")
                        
                        offers = []
                        for price_elem in price_elements:
                            try:
                                price_text = price_elem.text.strip()
                                print(f"[DEBUG] Armtek: проверяем цену: '{price_text}'")
                                
                                price_match = re.search(r'(\d[\d\s]*)\s*₽', price_text)
                                if price_match:
                                    price_value = float(price_match.group(1).replace(' ', ''))
                                    
                                    # Ищем количество рядом с ценой
                                    quantity = None
                                    try:
                                        # Ищем родительский контейнер предложения
                                        suggestion_item = price_elem.find_element(By.XPATH, "./ancestor::div[contains(@class, 'suggestion-item')]")
                                        qty_elements = suggestion_item.find_elements(By.CSS_SELECTOR, 'p.font__body2')
                                        
                                        for qty_elem in qty_elements:
                                            qty_text = qty_elem.text.strip()
                                            qty_match = re.search(r'(\d+)', qty_text)
                                            if qty_match:
                                                quantity = int(qty_match.group(1))
                                                print(f"[DEBUG] Armtek: найдено количество {quantity}")
                                                break
                                    except Exception as e:
                                        print(f"[DEBUG] Armtek: ошибка поиска количества: {str(e)}")
                                    
                                    if 100 <= price_value <= 100000:
                                        offers.append({
                                            'price': price_value,
                                            'quantity': quantity
                                        })
                                        print(f"[DEBUG] Armtek: добавлено предложение - цена: {price_value}₽, количество: {quantity}")
                            except Exception as e:
                                print(f"[DEBUG] Armtek: ошибка обработки цены: {str(e)}")
                                continue
                        
                        if offers:
                            # Сортируем по цене и берем минимальную
                            offers.sort(key=lambda x: x['price'])
                            best_offer = offers[0]
                            
                            result['marketplace_price'] = best_offer['price']
                            result['min_competitor_price'] = best_offer['price']
                            result['competitor_quantity'] = best_offer['quantity']
                            
                            print(f"[DEBUG] Armtek: выбрано лучшее предложение - цена: {best_offer['price']}₽, количество: {best_offer['quantity']}")
                        else:
                            print(f"[DEBUG] Armtek: предложения не найдены в секции 'Искомый товар'")
                            
                    except Exception as e:
                        print(f"[DEBUG] Armtek: ошибка поиска предложений в секции 'Искомый товар': {str(e)}")
                else:
                    print(f"[DEBUG] Armtek: бренд не совпадает - ожидали '{manufacturer}', нашли '{brand_text}'")
            else:
                print(f"[DEBUG] Armtek: товар не найден в секции 'Искомый товар'")
                
        except Exception as e:
            print(f"[DEBUG] Armtek: ошибка поиска в секции 'Искомый товар': {str(e)}")
        
    except Exception as e:
        result['error_message'] = f"Ошибка парсинга Армтек: {str(e)}"
        print(f"[DEBUG] Armtek: общая ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
    
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception:
                pass
    
    print(f"[DEBUG] Armtek: итоговый результат - найдено: {result['is_found']}, цена: {result['marketplace_price']}₽")
    return result

def create_result_excel(items: List[Dict], output_path: str) -> bool:
    """Создает Excel файл с результатами анализа"""
    print(f"[DEBUG] ===== НАЧАЛО create_result_excel =====")
    print(f"[DEBUG] Количество товаров для записи в Excel: {len(items)}")
    
    try:
        # Подготавливаем данные для Excel
        excel_data = []
        
        for i, item in enumerate(items, 1):
            # Логируем данные перед записью в Excel
            print(f"[DEBUG] Записываем в Excel - товар {i}:")
            print(f"[DEBUG]   - Цена наша: {item.get('marketplace_price')}")
            print(f"[DEBUG]   - Цена конкурента: {item.get('min_competitor_price')}")
            print(f"[DEBUG]   - Количество в наличии: {item.get('quantity_in_stock')}")
            print(f"[DEBUG]   - Количество конкурента: {item.get('competitor_quantity')}")
            print(f"[DEBUG]   - Тип quantity_in_stock: {type(item.get('quantity_in_stock'))}")
            print(f"[DEBUG]   - Тип competitor_quantity: {type(item.get('competitor_quantity'))}")
            
            # Получаем значения количества
            quantity_in_stock = item.get('quantity_in_stock')
            competitor_quantity = item.get('competitor_quantity')
            
            print(f"[DEBUG]   - quantity_in_stock после get: {quantity_in_stock}")
            print(f"[DEBUG]   - competitor_quantity после get: {competitor_quantity}")
            
            # Форматируем количество для Excel
            quantity_in_stock_str = f"{quantity_in_stock} шт" if quantity_in_stock is not None else ''
            competitor_quantity_str = f"{competitor_quantity} шт" if competitor_quantity is not None else ''
            
            print(f"[DEBUG]   - quantity_in_stock_str: '{quantity_in_stock_str}'")
            print(f"[DEBUG]   - competitor_quantity_str: '{competitor_quantity_str}'")
            
            row = {
                '№': i,
                'Бренд': item['manufacturer'],
                'Артикул по Бренду': item['article'],
                'Наименование': item['nomenclature'],
                'наличие': 'выгружено' if item['is_found'] else 'НЕТ',
                'источник': item.get('platform', ''),
                'Цена Наша': f"{item['marketplace_price']:.0f} ₽" if item['marketplace_price'] else '',
                'Минимальная цена конкурента': f"{item['min_competitor_price']:.0f} ₽" if item['min_competitor_price'] else '',
                'Количество в наличии': quantity_in_stock_str,
                'Количество конкурента': competitor_quantity_str
            }
            excel_data.append(row)
        
        # Создаем DataFrame и сохраняем в Excel
        df = pd.DataFrame(excel_data)
        
        # Настраиваем стили для Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Результаты', index=False)
            
            # Получаем рабочий лист для стилизации
            worksheet = writer.sheets['Результаты']
            
            # Применяем стили
            from openpyxl.styles import PatternFill, Font
            
            # Зеленый фон для "выгружено"
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            for row in range(2, len(excel_data) + 2):  # Начинаем с 2 (после заголовка)
                # Проверяем статус наличия
                status_cell = worksheet[f'E{row}']  # Колонка E - наличие
                if status_cell.value == 'выгружено':
                    status_cell.fill = green_fill
                elif status_cell.value == 'НЕТ':
                    status_cell.fill = red_fill
        
        print(f"[DEBUG] ===== КОНЕЦ create_result_excel =====")
        log_debug(f"Файл результата создан: {output_path}")
        return True
        
    except Exception as e:
        print(f"[DEBUG] ===== ОШИБКА create_result_excel =====")
        log_debug(f"Ошибка создания файла результата: {str(e)}")
        return False
